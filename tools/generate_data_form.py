#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generates content/data-to-fill.xlsx — a fill-in form listing every
placeholder / example / "준비 중" value currently shown on the site, so a
non-technical owner can supply the real data in one place. Each row:
위치(페이지) · 항목 · 현재 값 · [채워주실 값] · 비고.

    pip install -r tools/requirements-dev.txt   # openpyxl
    python3 tools/generate_data_form.py

This form is a MANUAL data-collection aid — it does not read/write the
site. Once returned filled, the values are applied by editing
tools/build_pages.py (code-managed items) or via /admin.html (news).
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "content", "data-to-fill.xlsx")

HEAD_FILL = PatternFill("solid", fgColor="06202B")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=15, color="06202B")
SUB_FONT = Font(size=10.5, color="445159")
PRIO_FILL = PatternFill("solid", fgColor="FBE9C7")   # 필수 강조
FILL_COL = PatternFill("solid", fgColor="EAF6F5")    # 입력 칸 하이라이트
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D6CE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ["위치 (페이지)", "항목", "현재 값 (플레이스홀더)", "✍️ 채워주실 값", "비고"]
WIDTHS = [22, 26, 40, 40, 30]


def sheet(wb, name, title, sub):
    ws = wb.create_sheet(name)
    ws["A1"] = title; ws["A1"].font = TITLE_FONT
    ws["A2"] = sub; ws["A2"].font = SUB_FONT
    ws.merge_cells("A1:E1"); ws.merge_cells("A2:E2")
    for i, c in enumerate(COLS, 1):
        cell = ws.cell(row=4, column=i, value=c)
        cell.font = HEAD_FONT; cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center"); cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
    ws.freeze_panes = "A5"; ws.row_dimensions[4].height = 22
    return ws


def rows(ws, data, start=5):
    r = start
    for loc, item, cur, note in data:
        ws.cell(row=r, column=1, value=loc).alignment = WRAP
        ws.cell(row=r, column=2, value=item).alignment = WRAP
        ws.cell(row=r, column=3, value=cur).alignment = WRAP
        fill = ws.cell(row=r, column=4); fill.alignment = WRAP; fill.fill = FILL_COL
        ws.cell(row=r, column=5, value=note).alignment = WRAP
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1
    return r


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook(); wb.remove(wb.active)

    # ---- 안내 ----
    g = wb.create_sheet("안내"); g.sheet_view.showGridLines = False
    g["A1"] = "컬리버 그룹 웹사이트 — 실제 데이터 입력 양식"
    g["A1"].font = Font(bold=True, size=16, color="06202B")
    lines = [
        "",
        "이 파일은 현재 사이트에 '예시 / 준비 중 / 000-0000-0000'처럼 임시로 들어가 있는 값을 한곳에 모은 목록입니다.",
        "",
        "사용법",
        "1. 각 탭의 '✍️ 채워주실 값'(하늘색 칸)에 실제 정보를 적어주세요. 모르거나 아직 없는 항목은 비워두시면 그대로 '준비 중'으로 남습니다.",
        "2. '현재 값'과 '항목' 열은 어디를 고치는지 찾는 기준이니 수정하지 마세요.",
        "3. 다 채운 파일을 전달해주시면 코드에 반영해 사이트를 다시 생성합니다.",
        "",
        "우선순위",
        "① 회사 기본정보 — 이메일·전화·주소·사업자등록번호는 지금 명백한 임시값(000-0000-0000 등)이라 가장 먼저 필요합니다.",
        "② 그룹 개요·연혁 — 설립연도, 각 계열사 합류 시점 등 사실관계 확인.",
        "③ 지표(계열사·지속가능·IR) — 실제 수치가 있으면 채우고, 없으면 비워두세요(자동으로 '준비 중' 유지).",
        "④ 채용 공고 — 현재 4개는 예시입니다. 실제 공고로 교체하거나, 채용 중이 아니면 알려주세요(섹션 처리 방식 상의).",
        "",
        "뉴스는 이 파일이 아닙니다",
        "뉴스룸 기사(보도자료·소식)는 /admin.html 관리자 페이지에서 직접 작성·수정·삭제합니다. 이 양식에는 포함하지 않았습니다.",
        "",
        "이미지",
        "대표이사 사진, 계열사/사업 이미지 등을 교체하려면 파일을 함께 전달해주세요(뉴스 이미지는 admin에서 업로드).",
    ]
    for i, ln in enumerate(lines, 3):
        cell = g.cell(row=i, column=1, value=ln)
        cell.font = Font(bold=ln in ("사용법", "우선순위", "뉴스는 이 파일이 아닙니다", "이미지"), size=11)
        cell.alignment = Alignment(wrap_text=True)
    g.column_dimensions["A"].width = 118

    # ---- ① 회사 기본정보 ----
    ws = sheet(wb, "① 회사 기본정보", "① 회사 기본정보 (필수)",
               "푸터·문의 페이지 전역에 공통으로 노출됩니다. 가장 먼저 채워주세요.")
    for c in range(1, 6):
        ws.cell(row=4, column=c).fill = PRIO_FILL if False else HEAD_FILL
    rows(ws, [
        ("푸터 · 문의", "대표 이메일", "contact@culiver.co.kr", "실제 대표 이메일"),
        ("푸터 · 문의", "대표 전화", "000-0000-0000", "실제 대표 전화번호"),
        ("푸터 · 문의 · 그룹개요", "본사 주소", "본사 주소를 입력하세요 (임시)", "도로명 주소 전체"),
        ("푸터 하단", "사업자 정보", "사업자 정보 · 주소는 실제 정보로 교체 (임시)", "법인명 / 사업자등록번호 / 대표자 등"),
        ("그룹소개(인사말)", "대표이사 성명", "'컬리버 그룹 대표이사' (이름 없음)", "대표이사 성함 (서명란에 표기)"),
        ("문의 폼 수신", "문의 접수 이메일", "환경변수 CONTACT_TO_EMAIL 미설정", "폼 제출을 받을 이메일 (Vercel 환경변수)"),
    ])

    # ---- ② 그룹 개요·연혁 ----
    ws = sheet(wb, "② 그룹 개요·연혁", "② 그룹 개요 · 연혁",
               "그룹소개(About) 페이지의 개요 표와 연혁 타임라인. 사실관계를 확인해주세요.")
    rows(ws, [
        ("그룹개요 표", "그룹명", "컬리버 그룹 · CULIVER GROUP", "확인 또는 수정"),
        ("그룹개요 표 · 홈 통계", "설립 연도", "2019년 (2026년 지주 체제 전환)", "실제 설립·지주전환 연도"),
        ("홈 히어로 통계", "핵심 숫자 4개", "계열사 4 / 무항생제 100% / 연중생산 365 / 시작 2019", "노출할 대표 지표 확정 (교체 가능)"),
        ("연혁", "2019 — 컬리버 설립", "육상 흰다리새우 BFT 양식 연구로 출발", "실제 연도·사건 내용"),
        ("연혁", "2021 — 에이엠피 합류", "수처리·미생물 기술 내재화", "실제 연도·내용"),
        ("연혁", "2023 — 코발티브 출범", "굴 패각 업사이클 사업 시작", "실제 연도·내용"),
        ("연혁", "2025 — 수신제팜 편입", "수경재배 스마트팜·신선유통 추가", "실제 연도·내용"),
        ("연혁", "2026 — 지주 체제 전환", "네 사업을 지주 체제로 정렬", "실제 연도·내용"),
    ])

    # ---- ③ 지표 ----
    ws = sheet(wb, "③ 지표(계열사·ESG·IR)", "③ 지표 (계열사 · 지속가능경영 · IR)",
               "‘예시 / 준비 중’으로 표시된 수치. 실제 값이 있으면 채우고, 없으면 비워두세요(그대로 '준비 중' 유지).")
    rows(ws, [
        ("에이엠피 상세", "누적 수주", "예시", "예: 120건 / 연 30건 등"),
        ("에이엠피 상세", "용수 재이용률(%)", "예시", "예: 85% 등"),
        ("코발티브 상세", "재활용 패각량", "예시", "예: 연 500톤 등"),
        ("코발티브 상세", "친환경 인증", "예시", "예: 환경표지 인증 등"),
        ("수신제팜 상세", "재배 면적", "예시", "예: 3,000㎡ 등"),
        ("수신제팜 상세", "연간 출하량", "예시", "예: 연 120톤 등"),
        ("지속가능경영", "양식수 재이용률", "예시", "실제 수치"),
        ("지속가능경영", "재활용 굴패각(톤)", "예시", "실제 수치"),
        ("지속가능경영", "지역 일자리", "예시", "실제 수치"),
        ("투자정보(IR)", "연결 매출", "준비 중", "실제 매출액 (공개 가능 시)"),
        ("투자정보(IR)", "영업이익", "준비 중", "실제 영업이익 (공개 가능 시)"),
        ("투자정보(IR)", "임직원 수", "준비 중", "실제 임직원 수"),
    ])

    # ---- ④ 자료·파일 ----
    ws = sheet(wb, "④ 자료·파일", "④ 다운로드 자료 (PDF 등)",
               "현재 '준비 중' 버튼으로 비활성화되어 있습니다. 파일을 주시면 실제 다운로드로 연결합니다.")
    rows(ws, [
        ("지속가능경영", "지속가능경영 보고서", "📄 (준비 중) 비활성", "PDF 파일 전달 시 연결"),
        ("투자정보(IR)", "감사보고서", "📄 (준비 중) 비활성", "PDF 파일 전달 시 연결"),
        ("투자정보(IR)", "IR 자료(IR Deck)", "📄 (준비 중) 비활성", "PDF/PPT 파일 전달 시 연결"),
    ])

    # ---- ⑤ 채용 공고 ----
    ws = sheet(wb, "⑤ 채용 공고", "⑤ 채용 공고 (현재 4개 모두 예시)",
               "실제 공고로 교체하거나, 채용 중이 아니면 알려주세요(안내 문구로 대체 가능).")
    rows(ws, [
        ("채용-컬리버", "직무 / 근무지", "양식 생산 매니저 / 충남 태안 / 정규직", "실제 직무·근무지·고용형태·직무기술서"),
        ("채용-에이엠피", "직무 / 근무지", "수처리 공정 엔지니어 / 경기 안산 / 정규직", "실제 직무·근무지·고용형태·직무기술서"),
        ("채용-코발티브", "직무 / 근무지", "소재 R&D 연구원 / 경남 통영 / 정규직", "실제 직무·근무지·고용형태·직무기술서"),
        ("채용-수신제팜", "직무 / 근무지", "스마트팜 재배 담당 / 전북 김제 / 정규직", "실제 직무·근무지·고용형태·직무기술서"),
        ("채용(공통)", "지원 방법", "지원하기 → 문의 폼 프리필", "이메일 지원/채용사이트 링크 등 원하는 방식"),
    ])

    # ---- ⑥ 계열사 소개 (선택) ----
    ws = sheet(wb, "⑥ 계열사 소개(선택)", "⑥ 계열사 상세 소개 (선택 보완)",
               "현재 소개 문구는 일반적 설명입니다. 더 정확한 내용이 있으면 보완해주세요.")
    rows(ws, [
        ("컬리버 상세", "소개·제품·기술", "BFT 흰다리새우 스마트양식 (일반 설명)", "실제 제품 라인업·기술 상세 (있으면)"),
        ("에이엠피 상세", "소개·제품·기술", "수처리·순환여과·미생물 제제 (일반 설명)", "실제 설비·수주 사례 (있으면)"),
        ("코발티브 상세", "소개·제품·기술", "굴패각 업사이클 '숨쉘'·'셸픽' (일반 설명)", "실제 제품·인증 (있으면)"),
        ("수신제팜 상세", "소개·제품·기술", "수경재배 스마트팜·신선유통 (일반 설명)", "실제 작물·유통 채널 (있으면)"),
    ])

    wb.save(OUT)
    print("saved", OUT)


if __name__ == "__main__":
    main()
