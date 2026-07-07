#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generates a per-page copy-editing spreadsheet (.xlsx) from the already-built
static HTML — one tab per page listing every bilingual (한글/영문) text
string, with blank "새 문구" columns for a non-technical editor to fill in.
Shared header/footer/mobile-menu text is split into one "공통" tab so it
isn't repeated 21 times.

Requires: beautifulsoup4, openpyxl (dev-only — not needed to run the site
itself, only to regenerate this spreadsheet).

    python3 tools/build_pages.py      # make sure HTML is up to date first
    python3 tools/generate_copy_sheet.py

Output: content/copy-editing-sheet.xlsx

Filled-in changes come back as data, not code — re-apply them to
tools/build_pages.py by hand (or hand the diff to Claude); this script
only extracts, it does not re-inject edits into the generator.
"""
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "content", "copy-editing-sheet.xlsx")

PAGES = [
    ("index.html", "홈"),
    ("about.html", "그룹소개"),
    ("business.html", "사업영역"),
    ("sustainability.html", "지속가능경영"),
    ("newsroom.html", "뉴스룸(목록)"),
    ("news.html", "뉴스 상세 템플릿"),
    ("careers.html", "채용(홈)"),
    ("careers-culiver.html", "채용-컬리버"),
    ("careers-amp.html", "채용-에이엠피"),
    ("careers-cobaltive.html", "채용-코발티브"),
    ("careers-susinje.html", "채용-수신제팜"),
    ("contact.html", "문의"),
    ("culiver-aqua.html", "컬리버 상세"),
    ("amp.html", "에이엠피 상세"),
    ("cobaltive.html", "코발티브 상세"),
    ("susinje-farm.html", "수신제팜 상세"),
]

HEADER_FILL = PatternFill("solid", fgColor="0B2438")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="0B2438")
NOTE_FONT = Font(size=10.5, color="445159")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D6CE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ["원문 (한글)", "원문 (영문)", "새 문구 (한글)", "새 문구 (영문)", "비고"]
COL_WIDTHS = [42, 42, 42, 42, 24]


def is_common(tag):
    return (
        tag.find_parent("header", class_="gnb") is not None
        or tag.find_parent("div", class_="mobile-menu") is not None
        or tag.find_parent("footer", class_="footer") is not None
    )


def extract_pairs(scope_tags):
    """scope_tags: iterable of class="t-ko" tags (h1/h2/p/span all carry this
    class, not just span) in document order. Returns deduped (ko, en) pairs,
    first-seen order, with occurrence counts."""
    seen = {}
    order = []
    for ko in scope_tags:
        en = ko.find_next_sibling()
        if not en or "t-en" not in (en.get("class") or []):
            continue
        ko_text = ko.get_text(" ", strip=True)
        en_text = en.get_text(" ", strip=True)
        if not ko_text and not en_text:
            continue
        key = (ko_text, en_text)
        if key not in seen:
            seen[key] = 0
            order.append(key)
        seen[key] += 1
    return [(ko, en, seen[(ko, en)]) for ko, en in order]


def style_sheet(ws: Worksheet, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    header_row = 4
    for i, col in enumerate(COLS, 1):
        c = ws.cell(row=header_row, column=i, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i - 1]
    ws.freeze_panes = f"A{header_row + 1}"
    ws.row_dimensions[header_row].height = 22
    return header_row


def write_rows(ws, start_row, rows):
    r = start_row
    for ko, en, count in rows:
        ws.cell(row=r, column=1, value=ko).alignment = WRAP
        ws.cell(row=r, column=2, value=en).alignment = WRAP
        ws.cell(row=r, column=3).alignment = WRAP
        ws.cell(row=r, column=4).alignment = WRAP
        note = "" if count <= 1 else f"페이지 내 {count}회 반복 (같은 문구, 한 번만 수정하면 됨)"
        ws.cell(row=r, column=5, value=note).alignment = WRAP
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    return r


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # ---------------------------------------------------------- 안내
    guide = wb.create_sheet("안내")
    guide.sheet_view.showGridLines = False
    guide["A1"] = "컬리버 그룹 웹사이트 — 페이지별 문구 변경 시트"
    guide["A1"].font = Font(bold=True, size=16, color="0B2438")
    lines = [
        "",
        "사용 방법",
        "1. 왼쪽 탭에서 문구를 바꾸고 싶은 페이지를 고르세요.",
        "2. '원문' 두 열(한글/영문)은 절대 수정하지 마세요 — 어떤 문구인지 찾는 기준입니다.",
        "3. 바꾸고 싶은 행에만 '새 문구 (한글)' / '새 문구 (영문)' 열을 채워주세요. 바꿀 필요 없는 행은 비워두면 됩니다.",
        "4. 다 채운 파일을 그대로 저장해서 담당자(또는 Claude)에게 전달하면, tools/build_pages.py에 반영해 사이트를 다시 생성합니다.",
        "",
        "주의할 점",
        "- '공통' 탭은 모든 페이지에 공통으로 나오는 상단 메뉴·하단 푸터·모바일 메뉴 문구입니다. 여기를 고치면 21개 페이지 전체에 한 번에 반영됩니다.",
        "- '비고' 열에 반복 횟수가 적힌 행은 같은 페이지 안에 같은 문구가 여러 번 나온다는 뜻입니다. 한 번만 새 문구를 적어도 전체에 반영됩니다.",
        "- 뉴스룸 기사(보도자료·소식·채용 소식)는 이 표가 아니라 /admin.html 관리자 페이지에서 직접 작성·수정·삭제합니다. '뉴스룸(목록)' / '뉴스 상세 템플릿' 탭에는 목록 화면의 안내 문구·필터 버튼 같은 틀만 들어 있습니다.",
        "- 채용 공고 문구는 이 표에서 바꿀 수 있습니다 (tools/build_pages.py의 정적 데이터).",
    ]
    for i, line in enumerate(lines, 3):
        cell = guide.cell(row=i, column=1, value=line)
        cell.font = Font(bold=line in ("사용 방법", "주의할 점"), size=11)
        cell.alignment = Alignment(wrap_text=True)
    guide.column_dimensions["A"].width = 110

    # ---------------------------------------------------------- 공통
    index_path = os.path.join(ROOT, "index.html")
    with open(index_path, encoding="utf-8") as f:
        index_soup = BeautifulSoup(f.read(), "html.parser")
    common_tags = [t for t in index_soup.find_all(class_="t-ko") if is_common(t)]
    common_rows = extract_pairs(common_tags)
    ws = wb.create_sheet("공통(헤더·푸터·메뉴)")
    hr = style_sheet(ws, "공통 — 헤더 / 모바일 메뉴 / 푸터", "모든 페이지에 공통으로 나오는 문구입니다. 여기를 고치면 21개 페이지 전체에 반영됩니다.")
    write_rows(ws, hr + 1, common_rows)
    print(f"공통: {len(common_rows)}개 문구")

    # ---------------------------------------------------------- per page
    for filename, display in PAGES:
        path = os.path.join(ROOT, filename)
        if not os.path.exists(path):
            print(f"skip (not found): {filename}")
            continue
        with open(path, encoding="utf-8") as f:
            soup = BeautifulSoup(f.read(), "html.parser")
        body_tags = [t for t in soup.find_all(class_="t-ko") if not is_common(t)]
        rows = extract_pairs(body_tags)
        sheet_name = display[:31]  # Excel tab-name limit
        ws = wb.create_sheet(sheet_name)
        hr = style_sheet(ws, f"{display} — {filename}", "이 페이지에만 나오는 문구입니다. '공통' 탭의 메뉴·푸터 문구는 여기 없습니다.")
        write_rows(ws, hr + 1, rows)
        print(f"{filename}: {len(rows)}개 문구")

    wb.save(OUT)
    print("saved", OUT)


if __name__ == "__main__":
    main()
